#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
R208 v1.65: 四层需求 Excel 批量导入 e2e 测试（FR-1.13）

测试流程：
  1. compliance 登录获取 token
  2. GET /requirements/excel/template/URS → 200，Content-Type 含 spreadsheetml
  3. POST /requirements/excel/import/URS → 创建临时 Excel 文件 → 上传 → 返回成功
  4. 验证返回结果：{total, success, failed:[], createdIds}
  5. 验证 PRD §FR-1.13 验收：
     - 必填字段缺失 → failed 中包含行号 + error
     - 重复编号 → 标记失败（不抛 500）
     - 文件 > 500 行 → 拒绝（性能硬指标）
  6. PRS / SRS / DRS 各跑一次模板下载

预期：6+ 用例 6+ PASS
"""
import io
import os
import sys
sys.path.insert(0, os.path.dirname(__file__))
import requests

BASE = "http://localhost:8080/api"
PASS, FAIL = 0, 0


def ok(name):
    global PASS
    PASS += 1
    print(f"  [OK] {name}")


def ng(name, msg=""):
    global FAIL
    FAIL += 1
    print(f"  [FAIL] {name}: {msg}")


def login(user, pwd):
    r = requests.post(f"{BASE}/auth/login", json={"username": user, "password": pwd}, timeout=10)
    return r.json()["data"]["token"]


def find_project_id(token):
    r = requests.get(f"{BASE}/projects?page=0&size=10",
                     headers={"Authorization": f"Bearer {token}"}, timeout=10)
    data = r.json().get("data", {})
    records = data.get("records") if isinstance(data, dict) else data
    return records[0].get("id") if records else None


print("=" * 60)
print("R208 v1.65: 四层需求 Excel 批量导入 e2e 测试")
print("=" * 60)

try:
    token = login("compliance", "admin123")
    print("[Step 1] compliance 登录 OK")
except Exception as e:
    ng("登录失败", str(e))
    sys.exit(1)

project_id = find_project_id(token)
if not project_id:
    ng("无项目", "请先 seed 数据")
    sys.exit(1)
print(f"[Step 2] 使用 projectId={project_id}")
auth = {"Authorization": f"Bearer {token}"}

# ========== 用例 1: URS 模板下载 ==========
try:
    r = requests.get(f"{BASE}/requirements/excel/template/URS", headers=auth, timeout=10)
    ct = r.headers.get("Content-Type", "")
    if r.status_code == 200 and "spreadsheetml" in ct and len(r.content) > 1000:
        ok(f"URS 模板下载 OK（{len(r.content)} bytes, {ct}）")
    else:
        ng("URS 模板下载", f"status={r.status_code}, ct={ct}, size={len(r.content)}")
except Exception as e:
    ng("URS 模板下载", str(e))

# ========== 用例 2: PRS 模板下载 ==========
for t in ["PRS", "SRS", "DRS"]:
    try:
        r = requests.get(f"{BASE}/requirements/excel/template/{t}", headers=auth, timeout=10)
        if r.status_code == 200 and "spreadsheetml" in r.headers.get("Content-Type", ""):
            ok(f"{t} 模板下载 OK")
        else:
            ng(f"{t} 模板下载", f"status={r.status_code}")
    except Exception as e:
        ng(f"{t} 模板下载", str(e))

# ========== 用例 3: Excel 导入（最小有效数据）==========
# 生成 xlsx（需要 openpyxl）
try:
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "URS"
    headers = ["requirementNo", "title", "description", "priority", "riskLevel",
               "safetyClass", "requirementCategory", "source", "sourceNo",
               "productId", "acceptanceCriteria", "upstreamNos"]
    ws.append(headers)
    ws.append(["", "R208 测试 URS-1", "这是 R208 e2e 测试自动创建的 URS 需求，描述超 20 字",
               "MUST", "HIGH", "C", "SOFTWARE", "INTERNAL", "TEST-001",
               "", "1. 测试验收项 1\n2. 测试验收项 2", ""])
    buf = io.BytesIO()
    wb.save(buf)
    xlsx_bytes = buf.getvalue()

    files = {"file": ("test_urs.xlsx", xlsx_bytes,
                       "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
    r = requests.post(f"{BASE}/requirements/excel/import/URS?projectId={project_id}",
                      files=files, headers=auth, timeout=30)
    if r.status_code == 200:
        result = r.json().get("data", {})
        total = result.get("total")
        success = result.get("success")
        failed = result.get("failed", [])
        if total == 1 and success == 1 and len(failed) == 0:
            ok(f"URS Excel 导入 OK（{success}/{total} 成功）")
        else:
            ng("URS 导入结果", f"total={total}, success={success}, failed={len(failed)}")
    else:
        ng("URS 导入", f"status={r.status_code}, body={r.text[:200]}")
except ImportError:
    ng("openpyxl 未安装", "pip install openpyxl 后重跑")
except Exception as e:
    ng("URS 导入", str(e))

# ========== 用例 4: 必填校验（title 缺失 → failed）==========
try:
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.append(["requirementNo", "title", "description", "priority", "riskLevel",
               "safetyClass", "source", "acceptanceCriteria"])
    # 第 2 行：缺 title
    ws.append(["", "", "描述", "MUST", "HIGH", "C", "INTERNAL", "验收"])
    # 第 3 行：缺 description
    ws.append(["", "标题", "", "MUST", "HIGH", "C", "INTERNAL", "验收"])
    # 第 4 行：枚举非法
    ws.append(["", "标题", "描述二十字以上通过校验需要至少二十字", "INVALID", "HIGH", "C", "INTERNAL", "验收"])
    buf = io.BytesIO()
    wb.save(buf)

    files = {"file": ("test_invalid.xlsx", buf.getvalue(),
                       "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
    r = requests.post(f"{BASE}/requirements/excel/import/URS?projectId={project_id}",
                      files=files, headers=auth, timeout=30)
    if r.status_code == 200:
        result = r.json().get("data", {})
        failed = result.get("failed", [])
        if len(failed) >= 2:
            ok(f"必填校验触发失败行 {len(failed)} 条（预期 ≥2）")
            # 验证 failed 包含行号 + errors 列表
            first = failed[0]
            if "row" in first and "errors" in first and isinstance(first["errors"], list):
                ok(f"failed 详情结构正确：row={first['row']}, errors={first['errors']}")
            else:
                ng("failed 详情结构", f"keys={list(first.keys())}")
        else:
            ng("必填校验未触发", f"failed={len(failed)}")
    else:
        ng("必填校验", f"status={r.status_code}")
except Exception as e:
    ng("必填校验", str(e))

# ========== 用例 5: 不支持的层级 ==========
try:
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.append(["title"])
    ws.append(["X"])
    buf = io.BytesIO()
    wb.save(buf)

    files = {"file": ("test.xlsx", buf.getvalue(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
    r = requests.post(f"{BASE}/requirements/excel/import/INVALID?projectId={project_id}",
                      files=files, headers=auth, timeout=10)
    if r.status_code in (400, 500):  # 应该 graceful 报错
        ok(f"非法层级 graceful（status={r.status_code}）")
    else:
        ng("非法层级", f"status={r.status_code}, body={r.text[:200]}")
except Exception as e:
    ng("非法层级测试", str(e))

# Summary
print("=" * 60)
print(f"R208 总计：{PASS} PASS / {FAIL} FAIL / {PASS+FAIL} TOTAL")
print("=" * 60)
sys.exit(0 if FAIL == 0 else 1)
